"""
Scraper mensual del IPC de Bolivia.

Fuentes:
  - INE (Instituto Nacional de Estadistica) - nube.ine.gob.bo
  - CEPALSTAT (CEPAL) - api-cepalstat.cepal.org

Salidas:
  data/ipc_general.json        - indice nacional + variaciones
  data/ipc_divisiones.json     - indice por division COICOP
  data/ipc_ciudades.json       - indice + variaciones por ciudad/conurbacion
  data/ipc_alimentos.json      - alimentos vs no alimentos (nacional)
  data/ipc_ciudades_alimentos.json - alimentos vs no alimentos por ciudad
  data/ipc_productos.json      - productos con mayor/menor variacion + ponderaciones
  data/ipc_productos_hist.json - series historicas de top productos
  data/ipc_ciudades_top.json   - top 5 subidas/bajadas por ciudad (para mapa)
  data/ipc_nucleo.json         - inflacion nucleo (exc. alimentos y transporte)
  data/ipc_transables.json     - IPC transables vs no transables (CEPALSTAT)
  data/ipc_regional.json       - IPC comparativo regional (CEPALSTAT)
  data/metadata.json           - fecha actualizacion, fuente, cobertura
"""

import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from io import BytesIO

import requests
import openpyxl

# ── Configuracion ────────────────────────────────────────────────────────────

DATA_DIR = Path(__file__).parent.parent / "data"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    )
}

FUENTES_NACIONAL = {
    "general":       "https://nube.ine.gob.bo/index.php/s/O4cCdvtUXQrhpNd/download",
    "divisiones":    "https://nube.ine.gob.bo/index.php/s/J4dSH7CTeHwL8SS/download",
    "productos":     "https://nube.ine.gob.bo/index.php/s/q3vO8P1pv65FGcm/download",
    "ponderaciones": "https://nube.ine.gob.bo/index.php/s/BjZRogQWdQBy2C0/download",
    "alimentos":     "https://nube.ine.gob.bo/index.php/s/vWT2R9HCIDsdjvE/download",
    "no_alimentos":  "https://nube.ine.gob.bo/index.php/s/yrJmRmHjmNXvw34/download",
}

FUENTES_CIUDADES = {
    "precios_promedio":      "https://nube.ine.gob.bo/index.php/s/crLd2jpbGZLiFAt/download",
    "ciudades_variaciones":  "https://nube.ine.gob.bo/index.php/s/zJiUiO88pgx3jSu/download",
    "ciudades_divisiones":   "https://nube.ine.gob.bo/index.php/s/VorZQVesKBFcEhE/download",
    "ciudades_alimentos":    "https://nube.ine.gob.bo/index.php/s/JSJsoJJNvcBPaBT/download",
    "ciudades_no_alimentos": "https://nube.ine.gob.bo/index.php/s/MS7cS0L4KSTCSjv/download",
    "ciudades_productos":    "https://nube.ine.gob.bo/index.php/s/ze8V3dJOhaHmf9n/download",
    "ciudades_ponderaciones":"https://nube.ine.gob.bo/index.php/s/M9PbTIGkUCgrP5z/download",
    "ponderacion_ciudades":  "https://nube.ine.gob.bo/index.php/s/n4repmm6wLInx89/download",
}

MESES = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4,
    "MAYO": 5, "JUNIO": 6, "JULIO": 7, "AGOSTO": 8,
    "SEPTIEMBRE": 9, "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

CIUDAD_DEPTO = {
    "BOLIVIA": "Nacional",
    "SUCRE": "Chuquisaca",
    "CONURBACION LA PAZ": "La Paz",
    "REGION METROPOLITANA KANATA": "Cochabamba",
    "ORURO": "Oruro",
    "POTOSI": "Potosi",
    "TARIJA": "Tarija",
    "CONURBACION SANTA CRUZ": "Santa Cruz",
    "TRINIDAD": "Beni",
    "COBIJA": "Pando",
}


def descargar(url: str) -> bytes:
    r = requests.get(url, headers=HEADERS, timeout=60, verify=False)
    r.raise_for_status()
    return r.content


def abrir_excel(contenido: bytes) -> openpyxl.Workbook:
    return openpyxl.load_workbook(BytesIO(contenido), data_only=True)


def safe_float(val) -> float | None:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def normalizar_ciudad(nombre: str) -> str:
    """Quita notas (1), (2), (3) y normaliza."""
    import re
    nombre = re.sub(r'\s*\(\d+\)\s*$', '', nombre).strip()
    return nombre


def depto_de_ciudad(ciudad: str) -> str:
    norm = ciudad.upper().replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    for key, val in CIUDAD_DEPTO.items():
        if key in norm:
            return val
    return ciudad


# ── Estructura tipo A: meses en filas, anos en columnas ─────────────────────
# Usado por: general, alimentos, no_alimentos
# R5: MES | 2018 | 2019 | 2020 | ...
# R7: Enero | val | val | val | ...

def leer_tipo_A(ws, fila_anios=5, fila_datos_inicio=7, col_nombre=1, col_datos_inicio=2):
    """
    Lee una hoja con meses en filas y anos en columnas.
    Retorna lista de {fecha: 'YYYY-MM', valor: float}.
    """
    # Leer anos del header
    anios = []
    for col in range(col_datos_inicio, ws.max_column + 1):
        val = ws.cell(row=fila_anios, column=col).value
        if val is not None:
            try:
                anios.append((col, int(val)))
            except (ValueError, TypeError):
                pass

    serie = []
    for fila in range(fila_datos_inicio, ws.max_row + 1):
        mes_raw = ws.cell(row=fila, column=col_nombre).value
        if mes_raw is None:
            continue
        mes_str = str(mes_raw).strip().upper()
        mes_num = MESES.get(mes_str)
        if mes_num is None:
            continue

        for col_anio, anio in anios:
            val = safe_float(ws.cell(row=fila, column=col_anio).value)
            if val is not None:
                serie.append({
                    "fecha": f"{anio}-{mes_num:02d}",
                    "valor": round(val, 6)
                })

    serie.sort(key=lambda x: x["fecha"])
    return serie


def leer_tipo_A_multifila(ws, fila_anios=5, fila_datos_inicio=7, col_nombre=1, col_datos_inicio=2):
    """
    Lee una hoja tipo A con multiples filas de datos (ciudades, divisiones).
    Retorna {nombre: [{fecha, valor}, ...]}.
    """
    anios = []
    for col in range(col_datos_inicio, ws.max_column + 1):
        val = ws.cell(row=fila_anios, column=col).value
        if val is not None:
            try:
                anios.append((col, int(val)))
            except (ValueError, TypeError):
                pass

    # Leer meses (fila_datos_inicio en adelante)
    # Pero primero necesitamos identificar los nombres de las filas
    resultado = {}
    for fila in range(fila_datos_inicio, ws.max_row + 1):
        nombre_raw = ws.cell(row=fila, column=col_nombre).value
        if nombre_raw is None:
            continue
        nombre = str(nombre_raw).strip()
        mes_num = MESES.get(nombre.upper())
        if mes_num is not None:
            continue
        if not nombre or nombre.upper() in ("MES", ""):
            continue

        serie = []
        for col_anio, anio in anios:
            val = safe_float(ws.cell(row=fila, column=col_anio).value)
            if val is not None:
                serie.append({"fecha": f"{anio}", "valor": round(val, 6)})

        if serie:
            resultado[nombre] = serie

    return resultado


# ── Estructura tipo B: ciudades en filas, meses en columnas ─────────────────
# Usado por: ciudades_variaciones
# R5: CIUDADES CAPITALES | 2018 | ... (merged)
# R6:                    | ENERO | FEBRERO | ...
# R8: BOLIVIA | val | val | ...

def leer_tipo_B(ws, fila_anios=5, fila_meses=6, fila_datos_inicio=8, col_nombre=1, col_datos_inicio=2):
    """
    Lee una hoja con ciudades en filas y meses en columnas.
    Retorna {ciudad: [{fecha, valor}, ...]}.
    """
    # Leer anos (se propagan a la derecha)
    anio_actual = None
    col_anios = {}
    for col in range(col_datos_inicio, ws.max_column + 1):
        val = ws.cell(row=fila_anios, column=col).value
        if val is not None:
            try:
                anio_actual = int(val)
            except (ValueError, TypeError):
                pass
        if anio_actual is not None:
            col_anios[col] = anio_actual

    # Leer meses
    col_fechas = {}
    for col in range(col_datos_inicio, ws.max_column + 1):
        val_mes = ws.cell(row=fila_meses, column=col).value
        if val_mes is None:
            continue
        mes_str = str(val_mes).strip().upper()
        mes_num = MESES.get(mes_str)
        if mes_num is None:
            continue
        anio = col_anios.get(col)
        if anio is not None:
            col_fechas[col] = f"{anio}-{mes_num:02d}"

    # Leer datos
    resultado = {}
    for fila in range(fila_datos_inicio, ws.max_row + 1):
        nombre_raw = ws.cell(row=fila, column=col_nombre).value
        if nombre_raw is None:
            continue
        nombre = normalizar_ciudad(str(nombre_raw).strip())
        if not nombre:
            continue

        serie = []
        for col, fecha in col_fechas.items():
            val = safe_float(ws.cell(row=fila, column=col).value)
            if val is not None:
                serie.append({"fecha": fecha, "valor": round(val, 6)})

        if serie:
            serie.sort(key=lambda x: x["fecha"])
            resultado[nombre] = serie

    return resultado


# ── Estructura tipo C: divisiones con codigo + descripcion ──────────────────
# R5: DIVISION | DESCRIPCION | 2018 | ... (merged)
# R6:          |             | ENERO | FEBRERO | ...
# R8: 0        | INDICE GENERAL | val | val | ...

def leer_tipo_C(ws, fila_anios=5, fila_meses=6, fila_datos_inicio=8,
                col_codigo=1, col_desc=2, col_datos_inicio=3):
    """
    Lee hoja con codigo + descripcion + datos mensuales.
    Retorna {descripcion: [{fecha, valor}, ...]}.
    """
    anio_actual = None
    col_anios = {}
    for col in range(col_datos_inicio, ws.max_column + 1):
        val = ws.cell(row=fila_anios, column=col).value
        if val is not None:
            try:
                anio_actual = int(val)
            except (ValueError, TypeError):
                pass
        if anio_actual is not None:
            col_anios[col] = anio_actual

    col_fechas = {}
    for col in range(col_datos_inicio, ws.max_column + 1):
        val_mes = ws.cell(row=fila_meses, column=col).value
        if val_mes is None:
            continue
        mes_str = str(val_mes).strip().upper()
        mes_num = MESES.get(mes_str)
        if mes_num is None:
            continue
        anio = col_anios.get(col)
        if anio is not None:
            col_fechas[col] = f"{anio}-{mes_num:02d}"

    resultado = {}
    for fila in range(fila_datos_inicio, ws.max_row + 1):
        desc_raw = ws.cell(row=fila, column=col_desc).value
        codigo_raw = ws.cell(row=fila, column=col_codigo).value
        if desc_raw is None:
            continue
        desc = str(desc_raw).strip()
        codigo = str(codigo_raw).strip() if codigo_raw is not None else ""
        if not desc:
            continue

        key = f"{codigo}. {desc}" if codigo and codigo != "0" else desc

        serie = []
        for col, fecha in col_fechas.items():
            val = safe_float(ws.cell(row=fila, column=col).value)
            if val is not None:
                serie.append({"fecha": fecha, "valor": round(val, 6)})

        if serie:
            serie.sort(key=lambda x: x["fecha"])
            resultado[key] = serie

    return resultado


# ── Procesadores por archivo ─────────────────────────────────────────────────

def procesar_general(contenido: bytes) -> dict:
    wb = abrir_excel(contenido)
    resultado = {}

    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper()
        if "INICIO" in upper:
            continue

        ws = wb[sheet_name]

        if "INDICE" in upper or "ÍNDICE" in upper:
            if "VAR" not in upper:
                resultado["indice"] = leer_tipo_A(ws)
        if "VAR" in upper and "MENSUAL" in upper and "ACUMULADA" not in upper and "12" not in upper:
            resultado["var_mensual"] = leer_tipo_A(ws)
        elif "ACUMULADA" in upper:
            resultado["var_acumulada"] = leer_tipo_A(ws)
        elif "12" in upper:
            resultado["var_interanual"] = leer_tipo_A(ws)

    return resultado


def procesar_divisiones(contenido: bytes) -> dict:
    wb = abrir_excel(contenido)
    resultado = {}

    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper()
        if "INICIO" in upper:
            continue
        if "INDICE" in upper or "ÍNDICE" in upper:
            if "VAR" not in upper:
                ws = wb[sheet_name]
                resultado = leer_tipo_C(ws)
                break

    return resultado


def procesar_ciudades(contenido: bytes) -> dict:
    wb = abrir_excel(contenido)
    resultado = {}

    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper()
        if "INICIO" in upper:
            continue

        ws = wb[sheet_name]

        if "INDICE" in upper or "ÍNDICE" in upper:
            if "VAR" not in upper:
                tipo = "indice"
            else:
                continue
        elif "MENSUAL" in upper and "ACUMULADA" not in upper and "12" not in upper:
            tipo = "var_mensual"
        elif "ACUMULADA" in upper:
            tipo = "var_acumulada"
        elif "12" in upper:
            tipo = "var_interanual"
        else:
            continue

        datos = leer_tipo_B(ws)
        for ciudad, serie in datos.items():
            if ciudad not in resultado:
                resultado[ciudad] = {"departamento": depto_de_ciudad(ciudad)}
            resultado[ciudad][tipo] = serie

    return resultado


def procesar_alimentos_simple(contenido: bytes) -> list:
    """Procesa un Excel de alimentos o no_alimentos (estructura tipo A)."""
    wb = abrir_excel(contenido)
    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper()
        if "INICIO" in upper:
            continue
        ws = wb[sheet_name]
        serie = leer_tipo_A(ws)
        if serie:
            return serie
    return []


def procesar_productos(contenido_prod: bytes, contenido_pond: bytes) -> dict:
    wb_pond = abrir_excel(contenido_pond)

    # Leer ponderaciones (estructura tipo C: codigo | descripcion | ponderacion)
    ponderaciones = {}
    for sheet_name in wb_pond.sheetnames:
        ws = wb_pond[sheet_name]
        if "INICIO" in sheet_name.upper():
            continue
        for fila in range(1, ws.max_row + 1):
            desc = ws.cell(row=fila, column=2).value
            pond = safe_float(ws.cell(row=fila, column=3).value)
            if pond is None:
                pond = safe_float(ws.cell(row=fila, column=2).value)
                desc = ws.cell(row=fila, column=1).value
            if desc and pond is not None:
                ponderaciones[str(desc).strip()] = pond

    # Leer indices de productos (estructura tipo C: codigo | descripcion | datos)
    wb_prod = abrir_excel(contenido_prod)
    productos = []

    for sheet_name in wb_prod.sheetnames:
        upper = sheet_name.upper()
        if "INICIO" in upper:
            continue

        ws = wb_prod[sheet_name]
        datos = leer_tipo_C(ws)

        for nombre, serie in datos.items():
            if len(serie) < 13:
                continue
            desc = nombre.split(". ", 1)[-1] if ". " in nombre else nombre
            ultimo = serie[-1]["valor"]
            hace_12 = serie[-13]["valor"] if len(serie) >= 13 else serie[0]["valor"]
            var_12 = ((ultimo / hace_12) - 1) * 100 if hace_12 else 0
            pond = ponderaciones.get(desc.strip(), None)
            productos.append({
                "producto": desc.strip(),
                "var_interanual": round(var_12, 4),
                "fecha": serie[-1]["fecha"],
                "ponderacion": pond,
            })
        break

    productos.sort(key=lambda x: x.get("var_interanual", 0), reverse=True)

    return {
        "top_subidas": productos[:15],
        "top_bajadas": list(reversed(productos[-15:])),
        "total_productos": len(productos),
    }


def procesar_ciudades_alimentos(contenido_ali: bytes, contenido_no_ali: bytes) -> dict:
    """Alimentos y no alimentos por ciudad. Cada ciudad es una hoja separada (tipo A)."""
    resultado = {}

    for label, contenido in [("alimentos", contenido_ali), ("no_alimentos", contenido_no_ali)]:
        wb = abrir_excel(contenido)
        for sheet_name in wb.sheetnames:
            upper = sheet_name.upper()
            if "INICIO" in upper:
                continue

            # Extraer nombre de ciudad del nombre de hoja (ej: "2 - SUCRE")
            parts = sheet_name.split(" - ", 1)
            ciudad = parts[-1].strip() if len(parts) > 1 else sheet_name.strip()
            ciudad = normalizar_ciudad(ciudad)

            ws = wb[sheet_name]
            serie = leer_tipo_A(ws)
            if serie:
                if ciudad not in resultado:
                    resultado[ciudad] = {"departamento": depto_de_ciudad(ciudad)}
                resultado[ciudad][label] = serie

    return resultado


def procesar_productos_historico(contenido_prod: bytes) -> dict:
    """Series historicas de los top 10 productos que mas subieron y bajaron."""
    wb = abrir_excel(contenido_prod)

    all_prods = {}
    for sheet_name in wb.sheetnames:
        if "INICIO" in sheet_name.upper():
            continue
        ws = wb[sheet_name]
        datos = leer_tipo_C(ws)
        for nombre, serie in datos.items():
            desc = nombre.split(". ", 1)[-1] if ". " in nombre else nombre
            all_prods[desc.strip()] = serie
        break

    # Calcular var interanual y seleccionar top/bottom 10
    ranked = []
    for nombre, serie in all_prods.items():
        if len(serie) < 13:
            continue
        ultimo = serie[-1]["valor"]
        hace_12 = serie[-13]["valor"]
        var_12 = ((ultimo / hace_12) - 1) * 100 if hace_12 else 0
        ranked.append((nombre, var_12))

    ranked.sort(key=lambda x: x[1], reverse=True)
    top_names = [r[0] for r in ranked[:10]] + [r[0] for r in ranked[-10:]]

    resultado = {}
    for nombre in top_names:
        if nombre in all_prods:
            resultado[nombre] = all_prods[nombre]

    return resultado


def procesar_ciudades_productos_top(contenido: bytes) -> dict:
    """Top 5 subidas y bajadas por ciudad (para tooltips del mapa)."""
    wb = abrir_excel(contenido)
    resultado = {}

    for sheet_name in wb.sheetnames:
        upper = sheet_name.upper()
        if "INICIO" in upper:
            continue

        ciudad = sheet_name.strip()
        ciudad = normalizar_ciudad(ciudad)

        ws = wb[sheet_name]
        datos = leer_tipo_C(ws)

        prods = []
        for nombre, serie in datos.items():
            if len(serie) < 13:
                continue
            desc = nombre.split(". ", 1)[-1] if ". " in nombre else nombre
            if "GENERAL" in desc.upper() or "INDICE" in desc.upper():
                continue
            ultimo = serie[-1]["valor"]
            hace_12 = serie[-13]["valor"]
            var_12 = ((ultimo / hace_12) - 1) * 100 if hace_12 else 0
            prods.append({"p": desc.strip(), "v": round(var_12, 2)})

        prods.sort(key=lambda x: x["v"], reverse=True)
        resultado[ciudad] = {
            "departamento": depto_de_ciudad(ciudad),
            "subidas": prods[:5],
            "bajadas": list(reversed(prods[-5:])),
        }

    return resultado


ENERGIA_CODIGOS = {"04510101", "04520101", "04520201", "04520202"}


def clasificar_producto(codigo: str) -> str:
    div = codigo[:2] if len(codigo) >= 2 else ""
    if div == "01":
        return "alimentos"
    if div == "07":
        return "energia"
    if codigo in ENERGIA_CODIGOS:
        return "energia"
    return "nucleo"


def calcular_descomposicion(contenido_prod: bytes, contenido_pond: bytes) -> dict:
    """
    Calcula indices ponderados para: general, nucleo, alimentos, energia.
    Usa datos a nivel de producto con ponderaciones reescaladas por categoria.
    """
    wb_pond = abrir_excel(contenido_pond)
    ponderaciones = {}
    for sheet_name in wb_pond.sheetnames:
        if "INICIO" in sheet_name.upper():
            continue
        ws = wb_pond[sheet_name]
        for fila in range(1, ws.max_row + 1):
            code = ws.cell(row=fila, column=1).value
            desc = ws.cell(row=fila, column=2).value
            pond = safe_float(ws.cell(row=fila, column=3).value)
            if code and desc and pond is not None:
                code_str = str(code).strip()
                if len(code_str) >= 4:
                    ponderaciones[code_str] = {
                        "desc": str(desc).strip(),
                        "pond": pond,
                        "cat": clasificar_producto(code_str),
                    }

    wb_prod = abrir_excel(contenido_prod)
    series_prod = {}
    for sheet_name in wb_prod.sheetnames:
        if "INICIO" in sheet_name.upper():
            continue
        ws = wb_prod[sheet_name]
        datos = leer_tipo_C(ws)
        for nombre, serie in datos.items():
            codigo = nombre.split(".")[0].strip()
            desc = nombre.split(". ", 1)[-1].strip() if ". " in nombre else nombre.strip()
            if codigo in ponderaciones:
                series_prod[codigo] = serie
            else:
                for pc, pi in ponderaciones.items():
                    if pi["desc"] == desc:
                        series_prod[pc] = serie
                        break
        break

    cats = {"general": {}, "nucleo": {}, "alimentos": {}, "energia": {}}
    for codigo, info in ponderaciones.items():
        if codigo not in series_prod:
            continue
        cat = info["cat"]
        cats["general"][codigo] = info["pond"]
        cats[cat][codigo] = info["pond"]

    resultado = {}
    for cat_name, prods_pond in cats.items():
        if not prods_pond:
            continue
        total_pond = sum(prods_pond.values())
        if total_pond == 0:
            continue
        pesos = {c: p / total_pond for c, p in prods_pond.items()}

        por_fecha = {}
        for codigo, peso in pesos.items():
            serie = series_prod.get(codigo, [])
            for punto in serie:
                f = punto["fecha"]
                if f not in por_fecha:
                    por_fecha[f] = 0.0
                por_fecha[f] += punto["valor"] * peso

        indice = [{"fecha": f, "valor": round(v, 6)}
                  for f, v in sorted(por_fecha.items())]
        resultado[cat_name] = indice

    meta = {}
    for cat_name, prods_pond in cats.items():
        total = sum(prods_pond.values())
        meta[cat_name] = {
            "productos": len(prods_pond),
            "ponderacion_total": round(total, 2),
        }
    resultado["meta"] = meta

    return resultado


def calcular_nucleo(divisiones: dict) -> list[dict]:
    """Fallback: promedio simple de divisiones excluyendo alimentos y transporte."""
    excluir_claves = {"1", "7", "01", "07"}
    excluir_nombres = {"ALIMENTOS", "TRANSPORTE"}

    divs_nucleo = {}
    for nombre, serie in divisiones.items():
        nombre_upper = nombre.upper()
        codigo = nombre.split(".")[0].strip()
        if codigo in excluir_claves:
            continue
        if any(e in nombre_upper for e in excluir_nombres):
            continue
        if "GENERAL" in nombre_upper or "INDICE GENERAL" in nombre_upper:
            continue
        divs_nucleo[nombre] = serie

    if not divs_nucleo:
        return []

    por_fecha = {}
    for serie in divs_nucleo.values():
        for punto in serie:
            f = punto["fecha"]
            if f not in por_fecha:
                por_fecha[f] = []
            por_fecha[f].append(punto["valor"])

    nucleo = []
    for f in sorted(por_fecha.keys()):
        vals = por_fecha[f]
        nucleo.append({"fecha": f, "valor": round(sum(vals) / len(vals), 6)})

    return nucleo


# ── CEPALSTAT ────────────────────────────────────────────────────────────────

CEPALSTAT_BASE = "https://api-cepalstat.cepal.org/cepalstat/api/v1"

CEPALSTAT_MES_MAP = {
    516: 1, 517: 2, 518: 3, 519: 4, 825: 5, 821: 6,
    822: 7, 823: 8, 824: 9, 826: 10, 827: 11, 828: 12,
}

CEPALSTAT_PAISES = {
    "BOL": "Bolivia", "ARG": "Argentina", "BRA": "Brasil",
    "CHL": "Chile", "COL": "Colombia", "ECU": "Ecuador",
    "PRY": "Paraguay", "PER": "Perú", "URY": "Uruguay",
}


def cepalstat_año(dim_id: int) -> int:
    return 1980 + (dim_id - 29150)


def cepalstat_fetch(indicator_id: int) -> list[dict]:
    url = f"{CEPALSTAT_BASE}/indicator/{indicator_id}/data?format=json&in=1&lang=es"
    r = requests.get(url, headers=HEADERS, timeout=120)
    r.raise_for_status()
    body = r.json().get("body", {})
    return body.get("data", [])


def cepalstat_to_series(raw: list[dict], filter_iso3: set | None = None) -> dict[str, list[dict]]:
    result = {}
    for rec in raw:
        iso3 = rec.get("iso3", "")
        if filter_iso3 and iso3 not in filter_iso3:
            continue
        val = safe_float(rec.get("value"))
        if val is None:
            continue
        dim_year = rec.get("dim_29117")
        dim_month = rec.get("dim_515")
        if dim_year is None or dim_month is None:
            continue
        year = cepalstat_año(dim_year)
        month = CEPALSTAT_MES_MAP.get(dim_month)
        if month is None or year < 2010:
            continue
        country = CEPALSTAT_PAISES.get(iso3, iso3)
        if country not in result:
            result[country] = []
        result[country].append({"fecha": f"{year}-{month:02d}", "valor": round(val, 4)})
    for series in result.values():
        series.sort(key=lambda x: x["fecha"])
    return result


def procesar_cepalstat_transables() -> dict:
    print("  > CEPALSTAT: IPC transables (762)...")
    raw_t = cepalstat_fetch(762)
    print("  > CEPALSTAT: IPC no transables (763)...")
    raw_nt = cepalstat_fetch(763)
    t = cepalstat_to_series(raw_t, {"BOL"})
    nt = cepalstat_to_series(raw_nt, {"BOL"})
    return {
        "transables": t.get("Bolivia", []),
        "no_transables": nt.get("Bolivia", []),
    }


def procesar_cepalstat_regional() -> dict:
    print("  > CEPALSTAT: IPC general regional (365)...")
    raw = cepalstat_fetch(365)
    return cepalstat_to_series(raw, set(CEPALSTAT_PAISES.keys()))


# ── Principal ────────────────────────────────────────────────────────────────

def main() -> None:
    import warnings
    warnings.filterwarnings("ignore", message="Unverified HTTPS")

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print("Descargando Excel del INE...")

    archivos = {}
    todas_fuentes = {**FUENTES_NACIONAL, **FUENTES_CIUDADES}

    for nombre, url in todas_fuentes.items():
        try:
            print(f"  > {nombre}...")
            archivos[nombre] = descargar(url)
        except Exception as e:
            print(f"  [ERROR] {nombre}: {e}", file=sys.stderr, flush=True)

    # ── Procesar nacional ────────────────────────────────────────────────
    print("\nProcesando datos nacionales...")

    if "general" in archivos:
        general = procesar_general(archivos["general"])
        out_path = DATA_DIR / "ipc_general.json"
        out_path.write_text(json.dumps(general, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        n = len(general.get("indice", []))
        print(f"  [OK] ipc_general.json - {n} meses")

    if "divisiones" in archivos:
        divisiones = procesar_divisiones(archivos["divisiones"])
        out_path = DATA_DIR / "ipc_divisiones.json"
        out_path.write_text(json.dumps(divisiones, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_divisiones.json - {len(divisiones)} divisiones")

        nucleo = calcular_nucleo(divisiones)
        if nucleo:
            nucleo_path = DATA_DIR / "ipc_nucleo.json"
            nucleo_path.write_text(json.dumps(nucleo, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
            print(f"  [OK] ipc_nucleo.json - {len(nucleo)} meses")

    if "alimentos" in archivos and "no_alimentos" in archivos:
        ali = procesar_alimentos_simple(archivos["alimentos"])
        no_ali = procesar_alimentos_simple(archivos["no_alimentos"])
        alimentos = {"alimentos": ali, "no_alimentos": no_ali}
        out_path = DATA_DIR / "ipc_alimentos.json"
        out_path.write_text(json.dumps(alimentos, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_alimentos.json - ali: {len(ali)}, no_ali: {len(no_ali)}")

    if "productos" in archivos and "ponderaciones" in archivos:
        productos = procesar_productos(archivos["productos"], archivos["ponderaciones"])
        out_path = DATA_DIR / "ipc_productos.json"
        out_path.write_text(json.dumps(productos, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_productos.json - {productos['total_productos']} productos")

        descomp = calcular_descomposicion(archivos["productos"], archivos["ponderaciones"])
        out_path = DATA_DIR / "ipc_descomposicion.json"
        out_path.write_text(json.dumps(descomp, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        meta = descomp.get("meta", {})
        for cat, info in meta.items():
            print(f"    {cat}: {info['productos']} prods, pond={info['ponderacion_total']}%")
        print(f"  [OK] ipc_descomposicion.json")

        if "nucleo" in descomp and descomp["nucleo"]:
            nucleo_path = DATA_DIR / "ipc_nucleo.json"
            nucleo_path.write_text(
                json.dumps(descomp["nucleo"], ensure_ascii=False, separators=(",", ":")),
                encoding="utf-8",
            )
            print(f"  [OK] ipc_nucleo.json (ponderado) - {len(descomp['nucleo'])} meses")

    if "productos" in archivos:
        prod_hist = procesar_productos_historico(archivos["productos"])
        out_path = DATA_DIR / "ipc_productos_hist.json"
        out_path.write_text(json.dumps(prod_hist, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_productos_hist.json - {len(prod_hist)} productos")

    # ── Procesar ciudades ────────────────────────────────────────────────
    print("\nProcesando datos por ciudad...")

    if "ciudades_variaciones" in archivos:
        ciudades = procesar_ciudades(archivos["ciudades_variaciones"])
        out_path = DATA_DIR / "ipc_ciudades.json"
        out_path.write_text(json.dumps(ciudades, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_ciudades.json - {len(ciudades)} ciudades")

    if "ciudades_alimentos" in archivos and "ciudades_no_alimentos" in archivos:
        ciudades_ali = procesar_ciudades_alimentos(
            archivos["ciudades_alimentos"], archivos["ciudades_no_alimentos"]
        )
        out_path = DATA_DIR / "ipc_ciudades_alimentos.json"
        out_path.write_text(json.dumps(ciudades_ali, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_ciudades_alimentos.json - {len(ciudades_ali)} ciudades")

    if "ciudades_productos" in archivos:
        ciudades_top = procesar_ciudades_productos_top(archivos["ciudades_productos"])
        out_path = DATA_DIR / "ipc_ciudades_top.json"
        out_path.write_text(json.dumps(ciudades_top, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_ciudades_top.json - {len(ciudades_top)} ciudades")

    # ── CEPALSTAT ────────────────────────────────────────────────────────
    print("\nDescargando datos CEPALSTAT...")
    try:
        transables = procesar_cepalstat_transables()
        out_path = DATA_DIR / "ipc_transables.json"
        out_path.write_text(json.dumps(transables, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_transables.json - T: {len(transables['transables'])}, NT: {len(transables['no_transables'])}")
    except Exception as e:
        print(f"  [ERROR] transables: {e}", file=sys.stderr, flush=True)

    try:
        regional = procesar_cepalstat_regional()
        out_path = DATA_DIR / "ipc_regional.json"
        out_path.write_text(json.dumps(regional, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_regional.json - {len(regional)} países")
    except Exception as e:
        print(f"  [ERROR] regional: {e}", file=sys.stderr, flush=True)

    # ── Metadata ─────────────────────────────────────────────────────────
    metadata = {
        "actualizado": datetime.now(timezone.utc).isoformat(),
        "fuente": "INE Bolivia - Indice de Precios al Consumidor (Base 2016)",
        "url_nacional": "https://www.ine.gob.bo/index.php/nacional/",
        "url_ciudades": "https://www.ine.gob.bo/index.php/ciudades-y-conurbaciones/",
        "cobertura": "9 ciudades capitales y conurbaciones",
        "frecuencia": "Mensual",
        "base": 2016,
    }
    meta_path = DATA_DIR / "metadata.json"
    meta_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[OK] metadata.json")
    print("Proceso completado!")


def main_cepalstat_only() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    print("Descargando solo datos CEPALSTAT...")
    try:
        transables = procesar_cepalstat_transables()
        out_path = DATA_DIR / "ipc_transables.json"
        out_path.write_text(json.dumps(transables, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_transables.json - T: {len(transables['transables'])}, NT: {len(transables['no_transables'])}")
    except Exception as e:
        print(f"  [ERROR] transables: {e}", file=sys.stderr, flush=True)
    try:
        regional = procesar_cepalstat_regional()
        out_path = DATA_DIR / "ipc_regional.json"
        out_path.write_text(json.dumps(regional, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
        print(f"  [OK] ipc_regional.json - {len(regional)} países")
    except Exception as e:
        print(f"  [ERROR] regional: {e}", file=sys.stderr, flush=True)
    print("CEPALSTAT completado!")


if __name__ == "__main__":
    if "--cepalstat-only" in sys.argv:
        main_cepalstat_only()
    else:
        main()
